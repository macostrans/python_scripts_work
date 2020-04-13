import os
import openpyxl
import pandas as pd
import wget
from gspread_pandas import Spread, Client
import sys

sheet_link = sys.argv[1]
sheet_name = sys.argv[2]
parameters_dict = {}
parameters_list = []
file = wget.download(sheet_link)
book = openpyxl.load_workbook(file)
sheet = book[sheet_name]
for i in range(2, sheet.max_row + 1):  # to get rows
    parameters_dict = {}
    for j in range(1, sheet.max_column + 1):  # to get columns
        parameters_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    parameters_list.append(parameters_dict)
os.remove(os.path.join(os.getcwd(), file))
lists = parameters_list
for each_list in lists:
    spread = Spread(each_list["gsheet_workbook_name"])
    data = pd.read_excel(each_list["excel_location"], each_list["excel_sheet_name"])
    spread.clear_sheet(spread.get_sheet_dims(each_list["gsheet_sheet_name"])[0], spread.get_sheet_dims(each_list["gsheet_sheet_name"])[1], each_list["gsheet_sheet_name"])
    spread.df_to_sheet(data, index=False, sheet=each_list["gsheet_sheet_name"], start='A1', replace=True)