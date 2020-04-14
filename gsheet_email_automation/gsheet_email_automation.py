import os
import sys
import openpyxl
import wget
from email.mime.multipart import MIMEMultipart
import smtplib
from email.mime.base import MIMEBase
from email import encoders

sheet_link = sys.argv[1]
sheet_name = sys.argv[2]
parameters_dict = {}
parameters_list = []
file = wget.download(sheet_link)
book = openpyxl.load_workbook(file)
sheet = book[sheet_name]
for i in range(2, sheet.max_row + 1):  # to get rows
    parameters_dict = {}
    for j in range(1, sheet.max_column + 1):  # to get columns
        parameters_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    parameters_list.append(parameters_dict)
os.remove(os.path.join(os.getcwd(), file))
lists = parameters_list

#The mail addresses and password
sender_address = 'kannan@tricog.com'
sender_pass = "****"

for each_list in lists:
    receiver_address = each_list["receiver-address"]
    # Setup the MIME
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = each_list["subject"]  # The subject line
    file = wget.download(each_list["attachement-link"], out=each_list["attachment-filename"])
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(file, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=file)  # or
    # part.add_header('Content-Disposition', 'attachment; filename="attachthisfile.csv"')
    message.attach(part)
    session = smtplib.SMTP('smtp.gmail.com', 587)  # use gmail with port
    session.starttls()  # enable security
    session.login(sender_address, sender_pass)  # login with mail_id and password
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    os.remove(file)
    print('Mail Sent to {} with attachment {}'.format(receiver_address, each_list["attachment-filename"]))
session.quit()
