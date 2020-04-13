import os
import openpyxl
import psutil
import wget
import win32com.client
import subprocess
import sys

sheet_link = r"https://docs.google.com/spreadsheets/d/1gaPljv0deyjIsqOZFecv3IoHw1GrvEFzYgv8rTO-L4M/export?format=xlsx"
sheet_name = "Sheet1"
parameters_dict = {}
parameters_list = []
file = wget.download(sheet_link)
book = openpyxl.load_workbook(file)
sheet = book[sheet_name]
for i in range(2, sheet.max_row + 1):  # to get rows
    parameters_dict = {}
    for j in range(1, sheet.max_column + 1):  # to get columns
        parameters_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    parameters_list.append(parameters_dict)
os.remove(os.path.join(os.getcwd(), file))
lists = parameters_list
#lists = get_params.get_params_from_gsheet(r"https://docs.google.com/spreadsheets/d/1gaPljv0deyjIsqOZFecv3IoHw1GrvEFzYgv8rTO-L4M/export?format=xlsx", "Sheet1")
if "EXCEL.EXE" in (p.name() for p in psutil.process_iter()):
    os.system("TASKKILL /F /IM Excel.exe")
xlapp = win32com.client.DispatchEx("Excel.Application")
xlapp.DisplayAlerts = False
for each_list in lists:
    #print(each_list["excel_location"])
    wb = xlapp.workbooks.open(each_list["excel_location"])
    #wb = xlapp.workbooks.open(r"C:\Users\Kannan\Desktop\Book2.xlsx")
    wb.RefreshAll()
    wb.Save()
    wb.Close()
    wb = None
xlapp.Quit()
if "EXCEL.EXE" in (p.name() for p in psutil.process_iter()):
    os.system("TASKKILL /F /IM Excel.exe")

